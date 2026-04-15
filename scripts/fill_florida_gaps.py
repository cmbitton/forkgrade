#!/usr/bin/env python3
"""
Fill missing Florida inspections by scraping the DBPR portal.

The bulk CSV/XLSX files that the main importer uses are incomplete — the portal
has inspections that never appear in the bulk exports. This script:

  1. Downloads current CSVs to build license_num → license_id mapping
  2. Async-scrapes inspectionDates.asp for every FL restaurant (200 concurrent)
  3. Diffs portal dates against DB to find missing inspections
  4. Async-scrapes inspectionDetail.asp for each missing inspection (200 concurrent)
  5. Writes missing inspections + violations to DB

Usage:
    python3 scripts/fill_florida_gaps.py --dry-run
    python3 scripts/fill_florida_gaps.py
    python3 scripts/fill_florida_gaps.py --from-cache   # skip scraping, reuse cached records
"""

import asyncio
import aiohttp
import csv
import io
import json
import math
import os
import re
import sys
import time
import urllib.request
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

if not os.environ.get('DATABASE_URL'):
    from dotenv import load_dotenv
    load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

DATES_CONCURRENCY  = 100   # concurrent requests for inspectionDates pages
DETAIL_CONCURRENCY = 50    # concurrent requests for inspectionDetail pages (lower = fewer portal errors)
COMMIT_EVERY       = 500

CSV_URLS = [
    (f'https://www2.myfloridalicense.com/sto/file_download/extracts/{n}fdinspi.csv',
     f'district {n}')
    for n in range(1, 8)
]
XLSX_URLS = [
    ('https://www2.myfloridalicense.com/sto/file_download/hr/fdinspi_2223.xlsx',
     'FY 22-23'),
    ('https://www2.myfloridalicense.com/hr/inspections/fdinspi_2324.xlsx',
     'FY 23-24'),
    ('https://www2.myfloridalicense.com/hr/inspections/fdinspi_2425.xlsx',
     'FY 24-25'),
]

DATES_URL  = 'https://www.myfloridalicense.com/inspectionDates.asp?SID=&id={license_id}'
DETAIL_URL = 'https://www.myfloridalicense.com/inspectionDetail.asp?InspVisitID={visit_id}&id={license_id}'

# ── Regex ─────────────────────────────────────────────────────────────────────

_TAGS_RE      = re.compile(r'<[^>]+>')
_SPACES_RE    = re.compile(r'\s+')
_TR_RE        = re.compile(r'<tr[^>]*>(.*?)</tr>', re.DOTALL | re.IGNORECASE)
_TD_RE        = re.compile(r'<td[^>]*>(.*?)</td>', re.DOTALL | re.IGNORECASE)
_LINK_RE      = re.compile(
    r'InspVisitID=(\d+)&(?:amp;)?id=(\d+)',
    re.IGNORECASE,
)
_DATE_RE      = re.compile(r'(\d{2}/\d{2}/\d{4})')
_VIOL_CODE_RE = re.compile(r'^\d{1,3}[A-Za-z]?-\d{1,3}-\d{1,2}$')
_SEV_PREFIX_RE = re.compile(r'^(High Priority|Intermediate|Basic)\s*[-:]\s*', re.IGNORECASE)
_PORTAL_SEV   = {'high priority': 'critical', 'intermediate': 'major', 'basic': 'minor'}
_SEV_WEIGHT   = {'critical': 3, 'major': 2, 'minor': 1}


def _strip_html(text):
    return _SPACES_RE.sub(' ', _TAGS_RE.sub(' ', text)).strip()


def _parse_date(s):
    for fmt in ('%m/%d/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _result(score):
    if score >= 75:  return 'Pass'
    if score >= 55:  return 'Pass with Conditions'
    return 'Fail'

# ── Phase 1: Build license_num → license_id mapping ──────────────────────────

def _download(url, label=''):
    for attempt in range(3):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0 (compatible)'})
            with urllib.request.urlopen(req, timeout=180) as r:
                return r.read()
        except Exception:
            if attempt < 2:
                time.sleep(1)
    return None


def build_license_map():
    """Download CSV files and build license_num → license_id mapping."""
    mapping = {}

    for url, label in CSV_URLS:
        data = _download(url, label)
        if not data:
            print(f'  WARN: failed to download {label}', flush=True)
            continue
        text = data.decode('utf-8', errors='replace')
        reader = csv.DictReader(io.StringIO(text))
        reader.fieldnames = [f.strip() for f in (reader.fieldnames or [])]
        for row in reader:
            row = {k.strip(): v for k, v in row.items()}
            lic_num = (row.get('License Number') or row.get('LICENSE NUMBER')
                       or row.get('LICENSE_NUMBER') or row.get('License_Number') or '').strip()
            lic_id = (row.get('License ID') or row.get('LICENSE_ID')
                      or row.get('License_ID') or row.get('LIC_ID') or '').strip()
            if lic_num and lic_id:
                mapping[lic_num] = lic_id
        print(f'  {label}: {len(mapping):,} total mappings', flush=True)

    # Fallback: XLSX files for restaurants not in current CSVs
    need_xlsx = False
    if need_xlsx:
        for url, label in XLSX_URLS:
            data = _download(url, label)
            if not data or not data.startswith(b'PK\x03\x04'):
                continue
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(next(rows))]
            for raw in rows:
                d = dict(zip(headers, raw))
                lic_num = str(d.get('LICENSE_NUMBER', '') or '').strip()
                lic_id = str(d.get('LIC_ID', '') or '').strip()
                if lic_num and lic_id and lic_num not in mapping:
                    mapping[lic_num] = lic_id
            wb.close()
            print(f'  {label}: {len(mapping):,} total mappings', flush=True)

    return mapping

# ── Phase 2: Scrape inspectionDates pages ─────────────────────────────────────

def parse_dates_page(html):
    """Parse inspectionDates.asp → list of (visit_id, date, insp_type)."""
    results = []
    for tr_m in _TR_RE.finditer(html):
        tr_html = tr_m.group(1)
        link_m = _LINK_RE.search(tr_html)
        if not link_m:
            continue
        visit_id = link_m.group(1)

        cells = [_strip_html(td_m.group(1)) for td_m in _TD_RE.finditer(tr_html)]
        if len(cells) < 2:
            continue

        # First cell has the date (inside link), second has inspection type
        date_m = _DATE_RE.search(cells[0])
        if not date_m:
            continue
        insp_date = _parse_date(date_m.group(1))
        if not insp_date:
            continue
        insp_type = cells[1].strip() or 'Routine Inspection'

        results.append((visit_id, insp_date, insp_type))
    return results


async def fetch_dates_page(session, license_id, sem):
    """Fetch one inspectionDates.asp page. Returns (license_id, results | None)."""
    url = DATES_URL.format(license_id=license_id)
    async with sem:
        for attempt in range(2):
            try:
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as r:
                    if r.status >= 400:
                        if attempt < 1:
                            await asyncio.sleep(0.3)
                            continue
                        return license_id, None
                    html = await r.text(errors='replace')
                    return license_id, parse_dates_page(html)
            except Exception:
                if attempt < 1:
                    await asyncio.sleep(0.3)
        return license_id, None


# ── Phase 4: Scrape inspectionDetail pages ────────────────────────────────────

def parse_detail_page(html):
    """Parse violations from inspectionDetail.asp page."""
    violations = []
    for tr_m in _TR_RE.finditer(html):
        cells = [_strip_html(td_m.group(1)) for td_m in _TD_RE.finditer(tr_m.group(1))]
        if len(cells) < 3:
            continue
        viol_code = cells[0].strip()
        desc = cells[2].strip()
        if not _VIOL_CODE_RE.match(viol_code):
            continue
        if not _SEV_PREFIX_RE.match(desc):
            continue
        m = _SEV_PREFIX_RE.match(desc)
        sev = _PORTAL_SEV.get(m.group(1).lower(), 'minor') if m else 'minor'
        violations.append({'code': viol_code, 'desc': desc, 'severity': sev})
    return violations


async def fetch_detail_page(session, visit_id, license_id, sem):
    """Fetch one inspectionDetail.asp page. Returns (visit_id, violations | None)."""
    url = DETAIL_URL.format(visit_id=visit_id, license_id=license_id)
    async with sem:
        for attempt in range(5):
            try:
                async with session.get(url, timeout=aiohttp.ClientTimeout(total=30)) as r:
                    if r.status >= 400:
                        await asyncio.sleep(1 * (attempt + 1))
                        continue
                    html = await r.text(errors='replace')
                    return visit_id, parse_detail_page(html)
            except Exception:
                await asyncio.sleep(1 * (attempt + 1))
        return visit_id, None


# ── Main ──────────────────────────────────────────────────────────────────────

async def run_async(restaurants, license_map, known_pairs, known_insp_ids, dry_run):
    """
    restaurants: list of (restaurant_id, source_id/license_num)
    license_map: license_num → license_id
    known_pairs: {restaurant_id: set of inspection dates}
    known_insp_ids: set of all inspection source_ids in DB
    """
    # Map license_id → restaurant info for lookup after scraping
    lid_to_info = {}
    skipped_no_lid = 0
    for rid, lic_num in restaurants:
        lid = license_map.get(lic_num)
        if not lid:
            skipped_no_lid += 1
            continue
        lid_to_info[lid] = (rid, lic_num)

    print(f'\nPhase 2: Scraping inspectionDates for {len(lid_to_info):,} restaurants '
          f'({DATES_CONCURRENCY} concurrent)...', flush=True)
    if skipped_no_lid:
        print(f'  ({skipped_no_lid:,} restaurants skipped — no license_id mapping)', flush=True)

    # ── Phase 2: Fetch all inspectionDates pages ──────────────────────────────
    sem = asyncio.Semaphore(DATES_CONCURRENCY)
    connector = aiohttp.TCPConnector(limit=DATES_CONCURRENCY, ttl_dns_cache=300,
                                      enable_cleanup_closed=True)

    all_missing = []  # (restaurant_id, license_id, visit_id, insp_date, insp_type)
    checked = errors = 0
    t0 = time.time()

    async with aiohttp.ClientSession(connector=connector,
                                      headers={'User-Agent': 'Mozilla/5.0 (compatible)'}) as session:
        lids = list(lid_to_info.keys())
        tasks = [
            asyncio.ensure_future(fetch_dates_page(session, lid, sem))
            for lid in lids
        ]

        for coro in asyncio.as_completed(tasks):
            license_id, portal_entries = await coro
            checked += 1
            if portal_entries is None:
                errors += 1
            else:
                rid, lic_num = lid_to_info[license_id]
                db_pairs = known_pairs.get(rid, set())
                for visit_id, insp_date, insp_type in portal_entries:
                    if (insp_date, insp_type.strip()) not in db_pairs:
                        all_missing.append((rid, license_id, visit_id, insp_date, insp_type))

            if checked % 500 == 0:
                elapsed = time.time() - t0
                rate = checked / elapsed if elapsed > 0 else 0
                print(f'  [{checked:,}/{len(lids):,}] '
                      f'{len(all_missing):,} missing found, '
                      f'{errors:,} errors — {rate:.0f} req/s',
                      flush=True)

    elapsed = time.time() - t0
    print(f'\nPhase 2 done: {checked:,} restaurants checked in {elapsed:.0f}s, '
          f'{len(all_missing):,} missing inspections found\n', flush=True)

    if not all_missing:
        print('No missing inspections. Done.', flush=True)
        return []

    # ── Phase 3: Fetch detail pages for missing inspections ───────────────────
    print(f'Phase 3: Fetching {len(all_missing):,} inspection details '
          f'({DETAIL_CONCURRENCY} concurrent)...', flush=True)

    sem2 = asyncio.Semaphore(DETAIL_CONCURRENCY)
    connector2 = aiohttp.TCPConnector(limit=DETAIL_CONCURRENCY, ttl_dns_cache=300,
                                       enable_cleanup_closed=True)

    # Build lookup: visit_id → (rid, license_id, insp_date, insp_type)
    vid_to_info = {}
    for rid, license_id, visit_id, insp_date, insp_type in all_missing:
        vid_to_info[visit_id] = (rid, license_id, insp_date, insp_type)

    records = []
    fetched = detail_errors = 0
    t1 = time.time()

    async with aiohttp.ClientSession(connector=connector2,
                                      headers={'User-Agent': 'Mozilla/5.0 (compatible)'}) as session:
        visit_ids = list(vid_to_info.keys())
        tasks = [
            asyncio.ensure_future(
                fetch_detail_page(session, vid, vid_to_info[vid][1], sem2)
            )
            for vid in visit_ids
        ]

        for coro in asyncio.as_completed(tasks):
            visit_id, violations = await coro
            fetched += 1
            rid, license_id, insp_date, insp_type = vid_to_info[visit_id]
            if violations is None:
                detail_errors += 1
            else:
                risk = sum(_SEV_WEIGHT[v['severity']] for v in violations)
                score = round(100 * math.exp(-risk * 0.05))

                records.append({
                    'restaurant_id': rid,
                    'visit_id':      visit_id,
                    'insp_date':     insp_date,
                    'insp_type':     insp_type,
                    'violations':    violations,
                    'risk':          risk,
                    'score':         score,
                })

            if fetched % 500 == 0 or fetched == len(visit_ids):
                elapsed = time.time() - t1
                rate = fetched / elapsed if elapsed > 0 else 0
                print(f'  [{fetched:,}/{len(visit_ids):,}] '
                      f'{len(records):,} ok, {detail_errors:,} errors — '
                      f'{rate:.0f} req/s', flush=True)

    elapsed = time.time() - t1
    print(f'\nPhase 3 done: {len(records):,} inspections fetched in {elapsed:.0f}s\n', flush=True)

    return records


CACHE_FILE = Path(__file__).parent / '.florida_gap_cache.json'


def _save_cache(records):
    """Save scraped records to JSON so Phase 4 can resume if it crashes."""
    serializable = []
    for r in records:
        s = dict(r)
        s['insp_date'] = r['insp_date'].isoformat()
        serializable.append(s)
    CACHE_FILE.write_text(json.dumps(serializable))
    print(f'  Saved {len(records):,} records to cache ({CACHE_FILE.name})', flush=True)


def _load_cache():
    """Load records from JSON cache."""
    data = json.loads(CACHE_FILE.read_text())
    for r in data:
        r['insp_date'] = date.fromisoformat(r['insp_date'])
    return data


def main():
    dry_run = '--dry-run' in sys.argv
    from_cache = '--from-cache' in sys.argv

    from app import create_app
    from app.db import db
    from app.models.restaurant import Restaurant
    from app.models.inspection import Inspection
    from app.models.violation import Violation

    app = create_app()

    # ── Phase 1: Build license mapping ────────────────────────────────────────
    print('Phase 1: Building license_num → license_id mapping from CSVs...', flush=True)
    t0 = time.time()
    license_map = build_license_map()
    print(f'  {len(license_map):,} mappings built in {time.time() - t0:.0f}s\n', flush=True)

    with app.app_context():
        # Get all FL restaurants
        fl_restaurants = (
            db.session.query(Restaurant.id, Restaurant.source_id)
            .filter(Restaurant.region == 'florida')
            .filter(Restaurant.source_id.isnot(None))
            .all()
        )
        print(f'Florida restaurants in DB: {len(fl_restaurants):,}', flush=True)

        # Get all known inspection (date, type) pairs per restaurant
        known_rows = (
            db.session.query(Inspection.restaurant_id, Inspection.inspection_date,
                             Inspection.inspection_type)
            .filter(Inspection.region == 'florida')
            .all()
        )
        known_pairs = {}
        for rid, d, t in known_rows:
            known_pairs.setdefault(rid, set()).add((d, (t or '').strip()))
        print(f'Known inspections in DB: {len(known_rows):,}', flush=True)

        # Get all known inspection source_ids for dedup
        known_insp_ids = {
            row[0] for row in db.session.execute(
                db.text("SELECT source_id FROM inspections WHERE region = 'florida'")
            ).fetchall()
        }

        # Check how many restaurants have license_id mappings
        mapped = sum(1 for _, src in fl_restaurants if src in license_map)
        print(f'Restaurants with license_id mapping: {mapped:,}/{len(fl_restaurants):,}', flush=True)

        # If many are unmapped, download XLSX files too
        unmapped = len(fl_restaurants) - mapped
        if unmapped > 1000:
            print(f'\n  {unmapped:,} unmapped — downloading XLSX files for full coverage...', flush=True)
            for url, label in XLSX_URLS:
                data = _download(url, label)
                if not data or not data.startswith(b'PK\x03\x04'):
                    print(f'  WARN: {label} failed or invalid', flush=True)
                    continue
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                rows = ws.iter_rows(values_only=True)
                headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(next(rows))]
                added = 0
                for raw in rows:
                    d = dict(zip(headers, raw))
                    lic_num = str(d.get('LICENSE_NUMBER', '') or '').strip()
                    lic_id = str(d.get('LIC_ID', '') or '').strip()
                    if lic_num and lic_id and lic_num not in license_map:
                        license_map[lic_num] = lic_id
                        added += 1
                wb.close()
                print(f'  {label}: +{added:,} new mappings ({len(license_map):,} total)', flush=True)

            mapped = sum(1 for _, src in fl_restaurants if src in license_map)
            print(f'  Now mapped: {mapped:,}/{len(fl_restaurants):,}\n', flush=True)

        # ── Run async scraping (or load from cache) ────────────────────────────
        if from_cache and CACHE_FILE.exists():
            print(f'Loading {CACHE_FILE.name} (skipping Phases 2-3)...', flush=True)
            records = _load_cache()
            print(f'  Loaded {len(records):,} records from cache', flush=True)
        else:
            records = asyncio.run(run_async(
                fl_restaurants, license_map, known_pairs, known_insp_ids, dry_run,
            ))

        if not records:
            return

        # Save cache before any DB writes so we can resume with --from-cache
        _save_cache(records)

        if dry_run:
            print('--dry-run: showing first 50 missing inspections:')
            for rec in records[:50]:
                print(f"  restaurant_id={rec['restaurant_id']} "
                      f"{rec['insp_date']} {rec['insp_type']} "
                      f"score={rec['score']} viols={len(rec['violations'])}")
            print(f'\n--dry-run: {len(records)} inspections would be added.')
            return

        # ── Phase 4: Write to DB ──────────────────────────────────────────────
        # Reconnect — the DB connection has been idle for hours during scraping
        try:
            db.session.remove()
        except Exception:
            pass
        db.engine.dispose()
        print(f'Phase 4: Writing {len(records):,} inspections to DB (fresh connection)...',
              flush=True)

        # Build restaurant lookup in batches to avoid massive IN clause
        rid_set = {r['restaurant_id'] for r in records}
        rid_to_restaurant = {}
        rid_list = list(rid_set)
        BATCH = 500
        for i in range(0, len(rid_list), BATCH):
            batch = rid_list[i:i + BATCH]
            for r in Restaurant.query.filter(Restaurant.id.in_(batch)).all():
                rid_to_restaurant[r.id] = r
            if (i // BATCH) % 10 == 0:
                print(f'  loaded {min(i + BATCH, len(rid_list)):,}/{len(rid_list):,} restaurants...',
                      flush=True)

        written = 0
        for rec in records:
            visit_id = rec['visit_id']

            # Use visit_id as source_id (prefixed to avoid collision with insp_nums)
            source_id = f'portal-{visit_id}'
            if source_id in known_insp_ids:
                continue

            insp = Inspection(
                restaurant_id   = rec['restaurant_id'],
                inspection_date = rec['insp_date'],
                source_id       = source_id,
                inspection_type = rec['insp_type'],
                score           = rec['score'],
                risk_score      = rec['risk'],
                result          = _result(rec['score']),
                region          = 'florida',
            )
            db.session.add(insp)
            db.session.flush()

            for v in rec['violations']:
                db.session.add(Violation(
                    inspection_id     = insp.id,
                    violation_code    = v['code'] or None,
                    description       = v['desc'],
                    severity          = v['severity'],
                    corrected_on_site = False,
                ))

            # Update latest_inspection_date if newer
            restaurant = rid_to_restaurant.get(rec['restaurant_id'])
            if restaurant:
                old_latest = restaurant.latest_inspection_date
                if old_latest is None or rec['insp_date'] > old_latest:
                    restaurant.latest_inspection_date = rec['insp_date']
                    restaurant.ai_summary = None

            known_insp_ids.add(source_id)
            written += 1

            if written % COMMIT_EVERY == 0:
                db.session.commit()
                print(f'  {written:,}/{len(records):,} written...', flush=True)

        db.session.commit()
        print(f'\nDone. +{written:,} inspections written to DB.', flush=True)

        # Clean up cache file after successful write
        if CACHE_FILE.exists():
            CACHE_FILE.unlink()
            print('  Cache file removed.', flush=True)


if __name__ == '__main__':
    main()
